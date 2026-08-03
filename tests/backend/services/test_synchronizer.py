"""
Pruebas de integración de Synchronizer contra ficheros .xlsx reales
(en un directorio temporal, nunca contra storage/ del proyecto).

Se inyecta un ExcelTemplateManager falso que solo resuelve rutas dentro
de tmp_path, y se monkeypatchea BACKUP_DIR para que las copias de
seguridad tampoco toquen storage/backup/ real. El resto de servicios
(ExcelReader, ExcelWriter, ExcelFinder, ProductManager, Registry) son
las implementaciones reales del proyecto.
"""

from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from config.constants import (
    DAY_HEADER_ROW,
    FIRST_DAY_COLUMN,
    FIRST_PRODUCT_ROW,
    MONTH_EXTRACTION_COLUMN,
    PRODUCT_CODE_COLUMN,
    TOTAL_STOCK_COLUMN,
    VALUE_COLUMN,
)
from models.delivery import Delivery
from models.product import Product
from models.sales_point import SalesPoint
from services.registry import Registry
from services.synchronizer import Synchronizer


class _FakeTemplateManager:
    """Resuelve rutas de plantilla y mensual dentro de tmp_path."""

    def __init__(self, monthly_dir, templates_dir):
        self._monthly_dir = monthly_dir
        self._templates_dir = templates_dir

    def ensure_month(self, year, month):
        return self._monthly_dir

    def get_excel_path(self, sales_point, year, month):
        return self._monthly_dir / f"{sales_point}_Julio_{year}.xlsx"

    def get_template_path(self, sales_point):
        return self._templates_dir / f"{sales_point}.xlsx"


def _build_monthly_workbook(path):
    """Construye un Excel mensual mínimo con la misma forma que la plantilla real."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "extraccion"

    for day in range(1, 32):
        worksheet.cell(row=DAY_HEADER_ROW, column=FIRST_DAY_COLUMN + day - 1).value = day

    row = FIRST_PRODUCT_ROW
    worksheet.cell(row=row, column=PRODUCT_CODE_COLUMN).value = "1"
    worksheet.cell(row=row, column=2).value = "PRODUCTO A"
    worksheet.cell(row=row, column=3).value = 0
    worksheet.cell(row=row, column=4).value = "UNIDAD"
    worksheet.cell(row=row, column=5).value = 1.0
    worksheet.cell(
        row=row, column=MONTH_EXTRACTION_COLUMN
    ).value = f"=SUM(G{row}:AK{row})"
    worksheet.cell(
        row=row, column=TOTAL_STOCK_COLUMN
    ).value = f"=C{row}+{worksheet.cell(row=row, column=MONTH_EXTRACTION_COLUMN).coordinate}"
    worksheet.cell(row=row, column=VALUE_COLUMN).value = f"=E{row}*{TOTAL_STOCK_COLUMN}{row}"

    total_row = row + 1
    worksheet.cell(row=total_row, column=1).value = "TOTAL"
    worksheet.cell(
        row=total_row, column=MONTH_EXTRACTION_COLUMN
    ).value = f"=SUM({worksheet.cell(row=row, column=MONTH_EXTRACTION_COLUMN).coordinate}:{worksheet.cell(row=row, column=MONTH_EXTRACTION_COLUMN).coordinate})"

    workbook.save(path)
    return path


def _make_delivery(sales_point="Bar_Piscina", day=1, code="1", quantity=48.0):
    return Delivery(
        sales_point=SalesPoint(name=sales_point),
        delivery_date=date(2026, 7, day),
        products=[
            Product(
                code=code,
                name="PRODUCTO A",
                format="UNIDAD",
                price=1.0,
                quantity=quantity,
            )
        ],
    )


@pytest.fixture
def paths(tmp_path):
    monthly_dir = tmp_path / "monthly"
    templates_dir = tmp_path / "templates"
    monthly_dir.mkdir()
    templates_dir.mkdir()
    return monthly_dir, templates_dir


@pytest.fixture
def synchronizer(paths, tmp_path, monkeypatch):
    monkeypatch.setattr("services.synchronizer.BACKUP_DIR", tmp_path / "backup")
    monthly_dir, templates_dir = paths

    registry = Registry(
        registry_file=tmp_path / "registry" / "imported_deliveries.json",
        backup_directory=tmp_path / "registry_backup",
        acquire_lock=False,
    )

    return Synchronizer(
        registry=registry,
        template_manager=_FakeTemplateManager(monthly_dir, templates_dir),
    )


def _day_cell_value(monthly_path, day, row=FIRST_PRODUCT_ROW):
    workbook = load_workbook(monthly_path, data_only=False)
    worksheet = workbook["extraccion"]
    value = worksheet.cell(row=row, column=FIRST_DAY_COLUMN + day - 1).value
    workbook.close()
    return value


class TestSynchronizeExistingProduct:
    def test_writes_the_quantity_in_the_correct_day_column(self, synchronizer, paths):
        monthly_dir, _ = paths
        excel_path = monthly_dir / "Bar_Piscina_Julio_2026.xlsx"
        _build_monthly_workbook(excel_path)

        totals = synchronizer.run([_make_delivery(day=13, quantity=3024.0)])

        assert totals.error_deliveries == 0
        assert totals.synchronized_deliveries == 1
        assert _day_cell_value(excel_path, day=13) == 3024.0
        assert _day_cell_value(excel_path, day=12) is None

    def test_two_different_days_do_not_interfere(self, synchronizer, paths):
        monthly_dir, _ = paths
        excel_path = monthly_dir / "Bar_Piscina_Julio_2026.xlsx"
        _build_monthly_workbook(excel_path)

        synchronizer.run([_make_delivery(day=13, quantity=3024.0)])
        synchronizer.run([_make_delivery(day=29, quantity=1512.0)])

        assert _day_cell_value(excel_path, day=13) == 3024.0
        assert _day_cell_value(excel_path, day=29) == 1512.0


class TestIdempotency:
    def test_running_the_same_delivery_twice_does_not_double_the_quantity(
        self, synchronizer, paths
    ):
        # Dentro de la misma sesión, el Registry ya sabe que está
        # sincronizada: la segunda vez se omite (skipped), no se
        # "recupera" (eso es un caso distinto, ver el test siguiente).
        monthly_dir, _ = paths
        excel_path = monthly_dir / "Bar_Piscina_Julio_2026.xlsx"
        _build_monthly_workbook(excel_path)

        delivery = _make_delivery(day=13, quantity=3024.0)

        first = synchronizer.run([delivery])
        second = synchronizer.run([_make_delivery(day=13, quantity=3024.0)])

        assert first.synchronized_deliveries == 1
        assert second.synchronized_deliveries == 0
        assert second.skipped_deliveries == 1
        assert _day_cell_value(excel_path, day=13) == 3024.0

    def test_recovers_when_the_registry_lost_track_but_excel_already_has_it(
        self, paths, tmp_path, monkeypatch
    ):
        # Este es el escenario real que nos encontramos en producción: el
        # Registry se pierde/reinicia, pero el Excel mensual ya contiene
        # la marca __SYNC_STATE__ de una sincronización previa. El
        # programa debe reconocerlo y NO volver a sumar la cantidad.
        monkeypatch.setattr("services.synchronizer.BACKUP_DIR", tmp_path / "backup")
        monthly_dir, templates_dir = paths
        excel_path = monthly_dir / "Bar_Piscina_Julio_2026.xlsx"
        _build_monthly_workbook(excel_path)

        first_registry = Registry(
            registry_file=tmp_path / "registry" / "imported_deliveries.json",
            backup_directory=tmp_path / "registry_backup",
            acquire_lock=False,
        )
        first_synchronizer = Synchronizer(
            registry=first_registry,
            template_manager=_FakeTemplateManager(monthly_dir, templates_dir),
        )
        first_synchronizer.run([_make_delivery(day=13, quantity=3024.0)])

        # Un Registry completamente nuevo y vacío, como tras perder el JSON.
        fresh_registry = Registry(
            registry_file=tmp_path / "registry_fresh" / "imported_deliveries.json",
            backup_directory=tmp_path / "registry_fresh_backup",
            acquire_lock=False,
        )
        fresh_synchronizer = Synchronizer(
            registry=fresh_registry,
            template_manager=_FakeTemplateManager(monthly_dir, templates_dir),
        )

        totals = fresh_synchronizer.run([_make_delivery(day=13, quantity=3024.0)])

        assert totals.recovered_deliveries == 1
        assert totals.error_deliveries == 0
        # La cantidad no se duplica: sigue siendo la de la única entrega real.
        assert _day_cell_value(excel_path, day=13) == 3024.0

    def test_conflicting_delivery_for_the_same_key_is_reported_as_an_error(
        self, synchronizer, paths
    ):
        monthly_dir, _ = paths
        excel_path = monthly_dir / "Bar_Piscina_Julio_2026.xlsx"
        _build_monthly_workbook(excel_path)

        synchronizer.run([_make_delivery(day=13, quantity=3024.0)])
        conflicting = synchronizer.run(
            [_make_delivery(day=13, quantity=999.0)]
        )

        assert conflicting.synchronized_deliveries == 0
        assert conflicting.error_deliveries == 1
        # La cantidad original no debe alterarse ante un conflicto.
        assert _day_cell_value(excel_path, day=13) == 3024.0


class TestNewProductCreation:
    def test_creates_the_product_in_monthly_and_template_when_missing(
        self, synchronizer, paths
    ):
        monthly_dir, templates_dir = paths
        excel_path = monthly_dir / "Bar_Piscina_Julio_2026.xlsx"
        template_path = templates_dir / "Bar_Piscina.xlsx"
        _build_monthly_workbook(excel_path)
        _build_monthly_workbook(template_path)

        delivery = _make_delivery(day=5, code="999", quantity=10.0)

        totals = synchronizer.run([delivery])

        assert totals.error_deliveries == 0
        assert totals.created_in_month == 1
        assert totals.created_in_template == 1

        workbook = load_workbook(excel_path)
        worksheet = workbook["extraccion"]
        assert worksheet.cell(row=FIRST_PRODUCT_ROW + 1, column=1).value == "999"
        assert (
            worksheet.cell(
                row=FIRST_PRODUCT_ROW + 1, column=FIRST_DAY_COLUMN + 4
            ).value
            == 10.0
        )
        workbook.close()

        template_workbook = load_workbook(template_path)
        template_worksheet = template_workbook["extraccion"]
        assert (
            template_worksheet.cell(row=FIRST_PRODUCT_ROW + 1, column=1).value
            == "999"
        )
        template_workbook.close()


class TestValidation:
    def test_delivery_without_products_is_reported_as_an_error(self, synchronizer, paths):
        monthly_dir, _ = paths
        _build_monthly_workbook(monthly_dir / "Bar_Piscina_Julio_2026.xlsx")

        delivery = Delivery(
            sales_point=SalesPoint(name="Bar_Piscina"),
            delivery_date=date(2026, 7, 1),
            products=[],
        )

        totals = synchronizer.run([delivery])

        assert totals.error_deliveries == 1
        assert totals.synchronized_deliveries == 0

    def test_run_rejects_a_non_list_argument(self, synchronizer):
        with pytest.raises(ValueError):
            synchronizer.run(None)
