import pytest
from openpyxl import Workbook

from excel.finder import ExcelFinder


@pytest.fixture
def worksheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "extraccion"

    # Fila 7 y 8 son productos válidos; fila 9 está vacía (sin código);
    # fila 10 repite el código 4788 para probar que se conserva la primera
    # aparición, tal y como documenta ExcelFinder.build_product_index.
    worksheet.cell(row=7, column=1).value = 14279
    worksheet.cell(row=8, column=1).value = "4788"
    worksheet.cell(row=9, column=1).value = None
    worksheet.cell(row=10, column=1).value = 4788

    return worksheet


class TestBuildProductIndex:
    def test_indexes_valid_codes_by_row(self, worksheet):
        finder = ExcelFinder()

        index = finder.build_product_index(worksheet=worksheet)

        assert index["14279"] == 7
        assert index["4788"] == 8

    def test_keeps_first_occurrence_of_a_duplicated_code(self, worksheet):
        finder = ExcelFinder()

        index = finder.build_product_index(worksheet=worksheet)

        # La fila 10 repite el código de la fila 8: debe ganar la primera.
        assert index["4788"] == 8

    def test_ignores_rows_without_a_valid_code(self, worksheet):
        finder = ExcelFinder()

        index = finder.build_product_index(worksheet=worksheet)

        assert len(index) == 2

    def test_empty_sheet_produces_empty_index(self):
        workbook = Workbook()
        worksheet = workbook.active

        finder = ExcelFinder()
        index = finder.build_product_index(worksheet=worksheet)

        assert index == {}


class TestFindDayColumn:
    def test_day_one_maps_to_first_day_column(self):
        finder = ExcelFinder()

        assert finder.find_day_column(day=1) == 7

    def test_day_thirty_one_maps_to_last_day_column(self):
        finder = ExcelFinder()

        assert finder.find_day_column(day=31) == 37

    @pytest.mark.parametrize("day", [0, -1, 32, 100])
    def test_rejects_days_outside_the_valid_range(self, day):
        finder = ExcelFinder()

        with pytest.raises(ValueError):
            finder.find_day_column(day=day)

    def test_rejects_non_integer_day(self):
        finder = ExcelFinder()

        with pytest.raises(TypeError):
            finder.find_day_column(day=15.5)
