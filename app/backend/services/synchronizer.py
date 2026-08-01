"""
Proyecto:
    Iberostar Inventory Synchronizer

Archivo:
    synchronizer.py

Descripción:
    Servicio encargado de sincronizar las entregas importadas
    con los Excel mensuales y sus plantillas correspondientes.

    La sincronización incorpora:
        - Validación completa de cada entrega.
        - Escritura idempotente para evitar cantidades duplicadas.
        - Registro interno de entregas aplicadas en cada Excel mensual.
        - Guardado atómico de los libros modificados.
        - Copias de seguridad antes de sustituir archivos existentes.
        - Recuperación del Registry cuando el Excel ya fue actualizado.
"""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
import math
import os
from pathlib import Path
import shutil
import traceback
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config.constants import DAY_HEADER_ROW
from config.settings import BACKUP_DIR

from models.delivery import Delivery
from models.product import Product

from excel.finder import ExcelFinder
from excel.product_manager import ProductManager
from excel.reader import ExcelReader
from excel.writer import ExcelWriter

from services.excel_template_manager import ExcelTemplateManager
from services.registry import Registry
from utils.delivery_identity import build_delivery_key, build_payload_hash


@dataclass(slots=True)
class _DeliveryResult:
    """Resultado interno de una sincronización individual."""

    written: int = 0
    existing_in_month: int = 0
    created_in_month: int = 0
    created_in_template: int = 0
    recovered_from_excel: bool = False


@dataclass(slots=True)
class _SynchronizationTotals:
    """
    Contadores y mensajes de error acumulados del proceso completo.

    Se devuelve desde ``Synchronizer.run()`` para que el llamante
    (``main.py``) pueda combinarlo con el resumen del ``Importer`` y
    mostrar un único resumen final consolidado, en vez de que cada
    entrega imprima su propio progreso por consola.
    """

    synchronized_deliveries: int = 0
    recovered_deliveries: int = 0
    skipped_deliveries: int = 0
    error_deliveries: int = 0
    products_written: int = 0
    created_in_month: int = 0
    created_in_template: int = 0
    error_messages: list[str] = field(default_factory=list)


class Synchronizer:
    """
    Coordina la sincronización de entregas con los Excel mensuales.

    Cada Excel mensual conserva en una hoja interna la clave y la huella
    de las entregas ya aplicadas. Esto permite reintentar una ejecución
    sin volver a sumar cantidades cuando el Excel se guardó correctamente,
    pero el Registry no pudo actualizarse.
    """

    _SYNC_SHEET_NAME = "__SYNC_STATE__"
    _SYNC_HEADERS = (
        "delivery_key",
        "payload_hash",
        "applied_at_utc",
        "delivery_date",
        "sales_point",
        "product_count",
    )

    def __init__(
        self,
        registry: Registry,
        *,
        excel_reader: ExcelReader | None = None,
        excel_finder: ExcelFinder | None = None,
        excel_writer: ExcelWriter | None = None,
        template_manager: ExcelTemplateManager | None = None,
        product_manager: ProductManager | None = None,
    ) -> None:
        """
        Inicializa los servicios necesarios.

        Las dependencias opcionales facilitan las pruebas sin alterar
        la forma habitual de creación: ``Synchronizer(registry)``.
        """

        if registry is None:
            raise ValueError("El Registry es obligatorio.")

        self.registry = registry
        self.excel_reader = excel_reader or ExcelReader()
        self.excel_finder = excel_finder or ExcelFinder()
        self.excel_writer = excel_writer or ExcelWriter()
        self.template_manager = template_manager or ExcelTemplateManager()
        self.product_manager = product_manager or ProductManager()

    # ======================================================
    # PUBLIC
    # ======================================================

    def run(
        self,
        deliveries: list[Delivery],
    ) -> _SynchronizationTotals:
        """
        Ejecuta el proceso completo de sincronización.

        Una entrega con errores no detiene las demás. Los errores no se
        imprimen inmediatamente: se acumulan en los totales devueltos para
        que el llamante pueda incluirlos en el resumen final del proceso.

        Returns:
            Los contadores y mensajes de error acumulados durante la
            ejecución.
        """

        self._validate_deliveries_collection(deliveries)

        totals = _SynchronizationTotals()

        if not deliveries:
            return totals

        for delivery in deliveries:
            sales_point_name = self._safe_sales_point_name(delivery)
            delivery_date_text = self._safe_delivery_date_text(delivery)

            try:
                self._validate_delivery(delivery)

                if self.registry.is_synchronized(delivery):
                    totals.skipped_deliveries += 1
                    continue

                result = self._synchronize_delivery(delivery=delivery)

                totals.synchronized_deliveries += 1
                totals.products_written += result.written
                totals.created_in_month += result.created_in_month
                totals.created_in_template += result.created_in_template

                if result.recovered_from_excel:
                    totals.recovered_deliveries += 1

            except (FileNotFoundError, PermissionError, OSError, ValueError) as error:
                totals.error_deliveries += 1
                totals.error_messages.append(
                    f"{delivery_date_text} | {sales_point_name or 'DESCONOCIDO'} | "
                    f"{type(error).__name__}: {error}"
                )

            except Exception as error:  # Protección por entrega, sin ocultar la traza.
                totals.error_deliveries += 1
                totals.error_messages.append(
                    f"{delivery_date_text} | {sales_point_name or 'DESCONOCIDO'} | "
                    f"ERROR INESPERADO {type(error).__name__}: {error}\n"
                    f"{traceback.format_exc().rstrip()}"
                )

        return totals

    # ======================================================
    # SYNCHRONIZATION
    # ======================================================

    def _synchronize_delivery(
        self,
        delivery: Delivery,
    ) -> _DeliveryResult:
        """Sincroniza una única entrega y devuelve sus contadores."""

        workbook: Workbook | None = None
        template_workbook: Workbook | None = None

        sales_point_name = delivery.sales_point.name.strip()
        delivery_date = self._as_date(delivery.delivery_date)
        delivery_key = build_delivery_key(delivery)
        payload_hash = build_payload_hash(delivery)

        self._ensure_registered(delivery)

        try:
            self.template_manager.ensure_month(
                year=delivery_date.year,
                month=delivery_date.month,
            )

            excel_path = self.template_manager.get_excel_path(
                sales_point=sales_point_name,
                year=delivery_date.year,
                month=delivery_date.month,
            )

            workbook, worksheet = self.excel_reader.read(
                workbook_path=excel_path,
            )

            sync_sheet = self._get_or_create_sync_sheet(workbook)
            applied_hash = self._find_applied_delivery_hash(
                sync_sheet=sync_sheet,
                delivery_key=delivery_key,
            )

            if applied_hash is not None:
                return self._recover_already_applied_delivery(
                    delivery=delivery,
                    delivery_key=delivery_key,
                    expected_hash=payload_hash,
                    applied_hash=applied_hash,
                )

            product_index = self._build_month_product_index(worksheet)
            day = delivery_date.day
            day_column = self._validate_day_column(
                worksheet=worksheet,
                day=day,
            )

            result, new_products = self._write_products(
                worksheet=worksheet,
                delivery=delivery,
                product_index=product_index,
                day_column=day_column,
            )

            template_path: Path | None = None

            if new_products:
                template_path = self.template_manager.get_template_path(
                    sales_point=sales_point_name,
                )

                template_workbook, template_worksheet = self.excel_reader.read(
                    workbook_path=template_path,
                )

                result.created_in_template = self._update_template(
                    template_worksheet=template_worksheet,
                    new_products=new_products,
                )

            # La plantilla se guarda primero. Si después falla el mensual,
            # el reintento simplemente encontrará esos productos ya creados.
            if template_workbook is not None and template_path is not None:
                if result.created_in_template > 0:
                    self._create_backup(
                        source_path=template_path,
                        category="templates",
                    )
                    self._atomic_save_workbook(
                        workbook=template_workbook,
                        target_path=template_path,
                    )

            # La marca se introduce en el mismo libro y se guarda junto con
            # las cantidades. Por eso ambas operaciones quedan vinculadas.
            self._append_delivery_marker(
                sync_sheet=sync_sheet,
                delivery=delivery,
                delivery_key=delivery_key,
                payload_hash=payload_hash,
            )

            self._create_backup(source_path=excel_path, category="monthly")

            self._atomic_save_workbook(
                workbook=workbook,
                target_path=excel_path,
            )

            # Solo después de confirmar el Excel en disco se actualiza
            # el Registry. Si este guardado falla, el marcador del Excel
            # impedirá volver a sumar cantidades en el siguiente intento.
            self._mark_registry_synchronized(delivery)

            return result

        finally:
            if template_workbook is not None:
                template_workbook.close()

            if workbook is not None:
                workbook.close()

    def _recover_already_applied_delivery(
        self,
        delivery: Delivery,
        delivery_key: str,
        expected_hash: str,
        applied_hash: str,
    ) -> _DeliveryResult:
        """
        Repara el Registry sin volver a escribir una entrega que ya
        figura aplicada dentro del Excel mensual.
        """

        if applied_hash != expected_hash:
            raise ValueError(
                "El Excel mensual ya contiene una entrega con la misma "
                "fecha y punto de venta, pero sus productos no coinciden. "
                f"Identificador: {delivery_key}. No se ha modificado el Excel."
            )

        self._mark_registry_synchronized(delivery)

        return _DeliveryResult(recovered_from_excel=True)

    # ======================================================
    # PRODUCT WRITING
    # ======================================================

    def _build_month_product_index(
        self,
        worksheet: Worksheet,
    ) -> dict[str, int]:
        """Construye el índice código → fila del Excel mensual."""

        return self.excel_finder.build_product_index(
            worksheet=worksheet,
        )

    def _validate_day_column(
        self,
        worksheet: Worksheet,
        day: int,
    ) -> int:
        """Localiza la columna del día y valida su cabecera real."""

        day_column = self.excel_finder.find_day_column(day=day)

        day_header_value = worksheet.cell(
            row=DAY_HEADER_ROW,
            column=day_column,
        ).value

        if not self._day_header_matches(day_header_value, day):
            raise ValueError(
                f"La columna {day_column} no corresponde al día {day}. "
                f"Valor encontrado: {day_header_value!r}"
            )

        return day_column

    def _write_products(
        self,
        worksheet: Worksheet,
        delivery: Delivery,
        product_index: dict[str, int],
        day_column: int,
    ) -> tuple[_DeliveryResult, list[Product]]:
        """Busca, crea y escribe todos los productos de una entrega."""

        result = _DeliveryResult()
        new_products: list[Product] = []

        for product in delivery.products:
            row, created = self.product_manager.find_or_create(
                worksheet=worksheet,
                product_index=product_index,
                product=product,
            )

            if created:
                new_products.append(product)
                result.created_in_month += 1
            else:
                result.existing_in_month += 1

            self.excel_writer.write(
                worksheet=worksheet,
                row=row,
                column=day_column,
                quantity=product.quantity,
            )

            result.written += 1

        return result, new_products

    def _update_template(
        self,
        template_worksheet: Worksheet,
        new_products: list[Product],
    ) -> int:
        """Añade a la plantilla los productos creados en el mensual."""

        template_product_index = self.excel_finder.build_product_index(
            worksheet=template_worksheet,
        )

        created_count = 0

        for product in new_products:
            _, created = self.product_manager.find_or_create(
                worksheet=template_worksheet,
                product_index=template_product_index,
                product=product,
            )

            if created:
                created_count += 1

        return created_count

    # ======================================================
    # IDEMPOTENCY MARKERS
    # ======================================================

    def _get_or_create_sync_sheet(
        self,
        workbook: Workbook,
    ) -> Worksheet:
        """Obtiene o crea la hoja interna de control de sincronización."""

        if self._SYNC_SHEET_NAME not in workbook.sheetnames:
            sync_sheet = workbook.create_sheet(self._SYNC_SHEET_NAME)

            for column, header in enumerate(self._SYNC_HEADERS, start=1):
                sync_sheet.cell(row=1, column=column).value = header

            sync_sheet.sheet_state = "veryHidden"
            return sync_sheet

        sync_sheet = workbook[self._SYNC_SHEET_NAME]

        existing_headers = tuple(
            sync_sheet.cell(row=1, column=column).value
            for column in range(1, len(self._SYNC_HEADERS) + 1)
        )

        if existing_headers != self._SYNC_HEADERS:
            raise ValueError(
                f"La hoja interna '{self._SYNC_SHEET_NAME}' está dañada. "
                f"Cabeceras esperadas: {self._SYNC_HEADERS}. "
                f"Cabeceras encontradas: {existing_headers}."
            )

        sync_sheet.sheet_state = "veryHidden"
        return sync_sheet

    def _find_applied_delivery_hash(
        self,
        sync_sheet: Worksheet,
        delivery_key: str,
    ) -> str | None:
        """Busca la huella asociada a una entrega ya aplicada."""

        matches: list[str] = []

        for row in range(2, sync_sheet.max_row + 1):
            stored_key = sync_sheet.cell(row=row, column=1).value

            if stored_key is None or str(stored_key).strip() != delivery_key:
                continue

            stored_hash = sync_sheet.cell(row=row, column=2).value

            if stored_hash is None or not str(stored_hash).strip():
                raise ValueError(
                    f"La entrega '{delivery_key}' existe en la hoja interna, "
                    "pero no contiene una huella válida."
                )

            matches.append(str(stored_hash).strip())

        if len(matches) > 1:
            raise ValueError(
                f"La entrega '{delivery_key}' aparece duplicada en la hoja "
                f"interna '{self._SYNC_SHEET_NAME}'."
            )

        return matches[0] if matches else None

    def _append_delivery_marker(
        self,
        sync_sheet: Worksheet,
        delivery: Delivery,
        delivery_key: str,
        payload_hash: str,
    ) -> None:
        """Registra la entrega dentro del mismo Excel mensual."""

        if self._find_applied_delivery_hash(sync_sheet, delivery_key) is not None:
            raise ValueError(
                f"No se puede registrar dos veces la entrega '{delivery_key}'."
            )

        row = sync_sheet.max_row + 1
        delivery_date = self._as_date(delivery.delivery_date)

        values = (
            delivery_key,
            payload_hash,
            datetime.now(timezone.utc).isoformat(timespec="seconds"),
            delivery_date.isoformat(),
            delivery.sales_point.name.strip(),
            len(delivery.products),
        )

        for column, value in enumerate(values, start=1):
            sync_sheet.cell(row=row, column=column).value = value

        sync_sheet.sheet_state = "veryHidden"

    # ======================================================
    # REGISTRY
    # ======================================================

    def _ensure_registered(
        self,
        delivery: Delivery,
    ) -> None:
        """Registra y persiste una entrega antes de modificar Excel."""

        if self.registry.exists(delivery):
            return

        snapshot = deepcopy(self.registry.data)

        try:
            self.registry.register(delivery)
            self.registry.save()

            if not self.registry.exists(delivery):
                raise RuntimeError("El Registry no confirmó el registro de la entrega.")

        except Exception:
            self._restore_registry_snapshot(snapshot)
            raise

    def _mark_registry_synchronized(
        self,
        delivery: Delivery,
    ) -> None:
        """Actualiza el Registry y revierte su memoria si falla el guardado."""

        if not self.registry.exists(delivery):
            raise ValueError(
                "No se puede completar la sincronización porque la entrega "
                "no está registrada."
            )

        snapshot = deepcopy(self.registry.data)

        try:
            self.registry.mark_as_synchronized(delivery)

            if not self.registry.is_synchronized(delivery):
                raise RuntimeError("El Registry no marcó la entrega como sincronizada.")

            self.registry.save()

        except Exception:
            self._restore_registry_snapshot(snapshot)
            raise

    def _restore_registry_snapshot(
        self,
        snapshot: dict,
    ) -> None:
        """Restaura el estado en memoria del Registry."""

        current_data = self.registry.data
        current_data.clear()
        current_data.update(snapshot)

    # ======================================================
    # SAFE FILE WRITING
    # ======================================================

    # Cantidad de copias de seguridad que se conservan por archivo de
    # origen. Las más antiguas se eliminan automáticamente al superar
    # este número, ya que cada sincronización genera una copia nueva.
    _BACKUP_RETENTION_COUNT = 15

    def _create_backup(
        self,
        source_path: Path,
        category: str,
    ) -> Path:
        """Crea una copia de seguridad única antes de modificar un Excel."""

        source_path = Path(source_path)

        if not source_path.is_file():
            raise FileNotFoundError(
                f"No se puede crear el backup porque no existe: {source_path}"
            )

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S_%f")
        backup_directory = BACKUP_DIR / category
        backup_directory.mkdir(parents=True, exist_ok=True)

        backup_path = backup_directory / (
            f"{source_path.stem}_{timestamp}{source_path.suffix}"
        )

        shutil.copy2(source_path, backup_path)

        self._prune_old_backups(
            backup_directory=backup_directory,
            stem=source_path.stem,
            suffix=source_path.suffix,
        )

        return backup_path

    def _prune_old_backups(
        self,
        backup_directory: Path,
        stem: str,
        suffix: str,
    ) -> None:
        """Elimina los backups más antiguos de un archivo por encima del límite."""

        existing_backups = sorted(
            backup_directory.glob(f"{stem}_*{suffix}"),
            key=lambda path: path.name,
        )

        excess_count = len(existing_backups) - self._BACKUP_RETENTION_COUNT

        if excess_count <= 0:
            return

        for old_backup in existing_backups[:excess_count]:
            old_backup.unlink(missing_ok=True)

    def _atomic_save_workbook(
        self,
        workbook: Workbook,
        target_path: Path,
    ) -> None:
        """Guarda un libro en un temporal y sustituye el destino al final."""

        target_path = Path(target_path)
        target_path.parent.mkdir(parents=True, exist_ok=True)

        temporary_path = target_path.parent / (
            f".{target_path.stem}.{uuid4().hex}.tmp{target_path.suffix}"
        )

        try:
            workbook.save(temporary_path)

            if not temporary_path.is_file() or temporary_path.stat().st_size == 0:
                raise OSError(
                    f"El archivo temporal no se creó correctamente: {temporary_path}"
                )

            os.replace(temporary_path, target_path)

        finally:
            temporary_path.unlink(missing_ok=True)

    # ======================================================
    # VALIDATION
    # ======================================================

    def _validate_deliveries_collection(
        self,
        deliveries: list[Delivery],
    ) -> None:
        """Valida el contenedor recibido por run()."""

        if deliveries is None:
            raise ValueError("La lista de entregas no puede ser None.")

        if not isinstance(deliveries, list):
            raise ValueError("Las entregas deben recibirse dentro de una lista.")

    def _validate_delivery(
        self,
        delivery: Delivery,
    ) -> None:
        """Valida una entrega antes de abrir o modificar archivos."""

        if not isinstance(delivery, Delivery):
            raise ValueError(f"Objeto de entrega no válido: {type(delivery).__name__}.")

        sales_point_name = self._safe_sales_point_name(delivery)

        if not sales_point_name:
            raise ValueError("La entrega no contiene un punto de venta válido.")

        self._as_date(delivery.delivery_date)

        if not isinstance(delivery.products, list) or not delivery.products:
            raise ValueError("La entrega no contiene ningún producto.")

        seen_codes: set[str] = set()

        for position, product in enumerate(delivery.products, start=1):
            if not isinstance(product, Product):
                raise ValueError(
                    f"El producto {position} no es una instancia válida de Product."
                )

            if not isinstance(product.code, str):
                raise ValueError(f"El código del producto {position} debe ser texto.")

            product_code = product.code.strip()

            if not product_code or not product_code.isdigit():
                raise ValueError(
                    f"Código de producto no válido en la posición {position}: "
                    f"{product.code!r}. Debe contener únicamente dígitos."
                )

            if product_code in seen_codes:
                raise ValueError(
                    f"El código {product_code} aparece repetido dentro de la entrega."
                )

            seen_codes.add(product_code)

            self._require_non_empty_text(
                product.name,
                f"nombre del producto {product_code}",
            )
            self._require_non_empty_text(
                product.format,
                f"formato del producto {product_code}",
            )

            price = self._require_finite_number(
                product.price,
                f"precio del producto {product_code}",
            )
            quantity = self._require_finite_number(
                product.quantity,
                f"cantidad del producto {product_code}",
            )

            if price < 0:
                raise ValueError(
                    f"El precio del producto {product_code} no puede ser negativo."
                )

            if math.isclose(quantity, 0.0, abs_tol=1e-9):
                raise ValueError(
                    f"La cantidad del producto {product_code} no puede ser cero."
                )

    def _as_date(
        self,
        value: object,
    ) -> date:
        """Devuelve una fecha válida sin componente horario."""

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        raise ValueError("La entrega no contiene una fecha válida.")

    def _require_non_empty_text(
        self,
        value: object,
        field_name: str,
    ) -> str:
        """Valida y limpia un campo textual."""

        if value is None:
            raise ValueError(f"El campo {field_name} está vacío.")

        normalized_value = " ".join(str(value).strip().split())

        if not normalized_value:
            raise ValueError(f"El campo {field_name} está vacío.")

        return normalized_value

    def _require_finite_number(
        self,
        value: object,
        field_name: str,
    ) -> float:
        """Valida un número real y finito."""

        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError(f"El campo {field_name} debe ser numérico.")

        parsed_value = float(value)

        if not math.isfinite(parsed_value):
            raise ValueError(f"El campo {field_name} debe ser un número finito.")

        return parsed_value

    def _day_header_matches(
        self,
        value: object,
        expected_day: int,
    ) -> bool:
        """Compara la cabecera admitiendo enteros, floats y texto numérico."""

        if isinstance(value, bool) or value is None:
            return False

        if isinstance(value, (int, float)):
            numeric_value = float(value)
            return math.isfinite(numeric_value) and numeric_value == expected_day

        normalized_value = str(value).strip()

        try:
            numeric_value = float(normalized_value.replace(",", "."))
        except ValueError:
            return False

        return math.isfinite(numeric_value) and numeric_value == expected_day

    def _safe_sales_point_name(
        self,
        delivery: object,
    ) -> str:
        """Obtiene el punto de venta sin provocar errores de presentación."""

        sales_point = getattr(delivery, "sales_point", None)
        name = getattr(sales_point, "name", "")
        return name.strip() if isinstance(name, str) else ""

    def _safe_delivery_date_text(
        self,
        delivery: object,
    ) -> str:
        """Obtiene la fecha en texto sin provocar errores de presentación."""

        delivery_date = getattr(delivery, "delivery_date", None)

        if isinstance(delivery_date, (date, datetime)):
            return delivery_date.strftime("%d/%m/%Y")

        return "FECHA DESCONOCIDA"
