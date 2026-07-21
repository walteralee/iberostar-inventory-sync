"""
Proyecto:
    Iberostar Inventory Synchronizer

Archivo:
    product_manager.py

Descripción:
    Servicio encargado de gestionar los productos dentro
    del Excel.
"""

import re
from copy import copy

from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from config.constants import (
    FIRST_PRODUCT_ROW,
    PRODUCT_CODE_COLUMN,
    PRODUCT_NAME_COLUMN,
    CURRENT_STOCK_COLUMN,
    PRODUCT_FORMAT_COLUMN,
    PRODUCT_PRICE_COLUMN,
)
from models.product import Product


class ProductManager:
    """
    Gestiona la búsqueda y creación de productos
    dentro del Excel.
    """

    def _insert_product_row(
        self,
        worksheet: Worksheet,
    ) -> int:
        """
        Inserta una nueva fila justo antes de la fila de totales.

        También desplaza correctamente las celdas combinadas
        pertenecientes a la fila de totales.
        """

        total_row = self._find_total_row(
            worksheet,
        )

        merged_ranges_to_move: list[tuple[int, int, int, int]] = []

        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.min_row <= total_row <= merged_range.max_row:
                merged_ranges_to_move.append(
                    (
                        merged_range.min_row,
                        merged_range.min_col,
                        merged_range.max_row,
                        merged_range.max_col,
                    )
                )

                worksheet.unmerge_cells(
                    str(merged_range),
                )

        worksheet.insert_rows(
            total_row,
        )

        for (
            min_row,
            min_column,
            max_row,
            max_column,
        ) in merged_ranges_to_move:
            worksheet.merge_cells(
                start_row=min_row + 1,
                start_column=min_column,
                end_row=max_row + 1,
                end_column=max_column,
            )

        return total_row

    def _find_total_row(
        self,
        worksheet: Worksheet,
    ) -> int:
        """
        Localiza la fila de totales situada después
        de los productos.
        """

        last_product_row = FIRST_PRODUCT_ROW - 1

        for row in range(
            FIRST_PRODUCT_ROW,
            worksheet.max_row + 1,
        ):
            code_value = worksheet.cell(
                row=row,
                column=PRODUCT_CODE_COLUMN,
            ).value

            if self._is_product_code(code_value):
                last_product_row = row
                continue

            if row > last_product_row and self._row_contains_formulas(worksheet, row):
                return row

        return last_product_row + 1

    def _is_product_code(
        self,
        value: object,
    ) -> bool:
        """
        Comprueba si un valor puede representar un código de producto.

        Los códigos se consideran identificadores numéricos, pero se
        conservan como texto para evitar perder ceros iniciales.
        """

        if value is None:
            return False

        if isinstance(value, bool):
            return False

        if isinstance(value, int):
            return True

        if isinstance(value, float):
            return value.is_integer()

        code = str(value).strip()

        return bool(code) and code.isdigit()

    def _row_contains_formulas(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> bool:
        """
        Comprueba si una fila contiene alguna fórmula.
        """

        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=row,
                column=column,
            ).value

            if isinstance(value, str) and value.startswith("="):
                return True

        return False

    def _copy_row_format(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> None:
        """
        Copia el formato de la fila del producto anterior.
        """

        source_row = row - 1

        worksheet.row_dimensions[row].height = worksheet.row_dimensions[
            source_row
        ].height

        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            source = worksheet.cell(
                row=source_row,
                column=column,
            )

            target = worksheet.cell(
                row=row,
                column=column,
            )

            target._style = copy(
                source._style,
            )

            if source.has_style:
                target.number_format = source.number_format

    def find_or_create(
        self,
        worksheet: Worksheet,
        product_index: dict[str, int],
        product: Product,
    ) -> tuple[int, bool]:
        """
        Devuelve la fila del producto y si ha sido creado.

        Returns:
            tuple[int, bool]:
                - Número de fila.
                - True si el producto fue creado.
                - False si el producto ya existía.
        """

        product_code = product.code.strip()

        if not product_code:
            raise ValueError("El código del producto no puede estar vacío.")

        row = product_index.get(
            product_code,
        )

        # ==================================================
        # PRODUCTO EXISTENTE
        # ==================================================

        if row is not None:
            print(
                f"Código: {product_code:<8} | "
                f"EXISTENTE | "
                f"Fila: {row:<5} | "
                f"No se modifican sus datos"
            )

            return row, False

        # ==================================================
        # PRODUCTO NUEVO
        # ==================================================

        print(f"Código: {product_code:<8} | " f"NUEVO     | " f"Creando producto...")

        row = self._insert_product_row(
            worksheet,
        )

        print(f"{'':18}│ Fila insertada          : {row}")

        self._copy_row_format(
            worksheet,
            row,
        )

        print(f"{'':18}│ Formato copiado         : SÍ")

        self._fill_product_data(
            worksheet,
            row,
            product,
        )

        print(f"{'':18}│ Código                  : {product_code}")
        print(f"{'':18}│ Nombre                  : {product.name}")
        print(f"{'':18}│ Stock actual            : 0")
        print(f"{'':18}│ Formato                 : {product.format}")
        print(f"{'':18}│ Precio                  : {product.price}")

        self._copy_product_formulas(
            worksheet,
            row,
        )

        print(f"{'':18}│ Fórmulas adaptadas      : SÍ")

        self._update_total_formulas(
            worksheet,
            row,
        )

        print(f"{'':18}│ Totales actualizados    : SÍ")

        product_index[product_code] = row

        print(f"{'':18}│ Índice actualizado      : SÍ")
        print(f"{'':18}└─ Resultado               : PRODUCTO CREADO")

        return row, True

    def _fill_product_data(
        self,
        worksheet: Worksheet,
        row: int,
        product: Product,
    ) -> None:
        """
        Escribe los datos principales del nuevo producto.
        """

        code_cell = worksheet.cell(
            row=row,
            column=PRODUCT_CODE_COLUMN,
        )

        code_cell.value = product.code.strip()
        code_cell.number_format = "@"

        worksheet.cell(
            row=row,
            column=PRODUCT_NAME_COLUMN,
        ).value = product.name

        worksheet.cell(
            row=row,
            column=CURRENT_STOCK_COLUMN,
        ).value = 0

        worksheet.cell(
            row=row,
            column=PRODUCT_FORMAT_COLUMN,
        ).value = product.format

        worksheet.cell(
            row=row,
            column=PRODUCT_PRICE_COLUMN,
        ).value = product.price

    def _copy_product_formulas(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> None:
        """
        Copia las fórmulas de la fila anterior y adapta
        sus referencias a la nueva fila.
        """

        source_row = row - 1

        for column in range(
            6,
            worksheet.max_column + 1,
        ):
            source = worksheet.cell(
                row=source_row,
                column=column,
            )

            target = worksheet.cell(
                row=row,
                column=column,
            )

            if isinstance(source.value, str) and source.value.startswith("="):
                target.value = Translator(
                    source.value,
                    origin=source.coordinate,
                ).translate_formula(
                    target.coordinate,
                )

    def _update_total_formulas(
        self,
        worksheet: Worksheet,
        row: int,
    ) -> None:
        """
        Amplía las fórmulas de la fila de totales para incluir
        el producto recién insertado.
        """

        total_row = row + 1
        previous_last_product_row = row - 1

        range_end_pattern = re.compile(
            rf"(:\$?[A-Z]{{1,3}}\$?){previous_last_product_row}(?!\d)"
        )

        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            cell = worksheet.cell(
                row=total_row,
                column=column,
            )

            formula = cell.value

            if not (isinstance(formula, str) and formula.startswith("=")):
                continue

            cell.value = range_end_pattern.sub(
                lambda match: f"{match.group(1)}{row}",
                formula,
            )
