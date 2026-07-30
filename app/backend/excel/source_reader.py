"""
Proyecto:
    Iberostar Inventory Synchronizer

Archivo:
    source_reader.py

Descripción:
    Servicio encargado de abrir y proporcionar acceso
    al Excel de origen.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config.constants import (
    SOURCE_HEADER_ROW,
    SOURCE_DATE_COLUMN,
    SOURCE_SALES_POINT_COLUMN,
    SOURCE_GROUP_COLUMN,
    SOURCE_PRODUCT_CODE_COLUMN,
    SOURCE_PRODUCT_NAME_COLUMN,
    SOURCE_FORMAT_COLUMN,
    SOURCE_QUANTITY_COLUMN,
    SOURCE_PRICE_COLUMN,
)


class SourceReader:
    """
    Servicio encargado de abrir el Excel de origen.

    No depende de la hoja activa ni del nombre exacto de las
    cabeceras. Busca automáticamente la hoja que tenga las
    dimensiones necesarias y mayor cantidad de datos útiles.
    """

    _REQUIRED_COLUMNS = (
        SOURCE_DATE_COLUMN,
        SOURCE_SALES_POINT_COLUMN,
        SOURCE_GROUP_COLUMN,
        SOURCE_PRODUCT_CODE_COLUMN,
        SOURCE_PRODUCT_NAME_COLUMN,
        SOURCE_FORMAT_COLUMN,
        SOURCE_QUANTITY_COLUMN,
        SOURCE_PRICE_COLUMN,
    )

    def read(
        self,
        excel_file: Path,
    ) -> Worksheet:
        """
        Abre el Excel y devuelve la hoja de movimientos.

        Args:
            excel_file: Ruta del Excel de origen.

        Returns:
            Hoja seleccionada para procesar los movimientos.

        Raises:
            FileNotFoundError:
                Si el archivo no existe.

            ValueError:
                Si no se encuentra ninguna hoja con las
                dimensiones mínimas necesarias.
        """

        excel_file = Path(excel_file)

        if not excel_file.is_file():
            raise FileNotFoundError(f"No se encontró el Excel de origen: {excel_file}")

        workbook = load_workbook(
            filename=excel_file,
            data_only=True,
        )

        compatible_worksheets = [
            worksheet
            for worksheet in workbook.worksheets
            if self._is_valid_source_sheet(worksheet)
        ]

        if not compatible_worksheets:
            workbook.close()

            raise ValueError(
                "No se encontró ninguna hoja con las filas y columnas "
                "necesarias para importar los movimientos de Economato."
            )

        return max(
            compatible_worksheets,
            key=self._calculate_sheet_score,
        )

    # ======================================================
    # PRIVATE
    # ======================================================

    def _is_valid_source_sheet(
        self,
        worksheet: Worksheet,
    ) -> bool:
        """
        Comprueba únicamente que la hoja tenga las dimensiones
        mínimas necesarias.

        No valida el texto ni el contenido de las cabeceras,
        porque los informes de Economato pueden utilizar nombres
        diferentes o contener algunas cabeceras vacías.
        """

        required_last_column = max(self._REQUIRED_COLUMNS)

        return (
            worksheet.max_row > SOURCE_HEADER_ROW
            and worksheet.max_column >= required_last_column
        )

    def _calculate_sheet_score(
        self,
        worksheet: Worksheet,
    ) -> int:
        """
        Calcula cuántos datos existen en las columnas configuradas.

        Si el libro contiene varias hojas compatibles, se selecciona
        la que tenga mayor cantidad de información en las columnas
        utilizadas por el Importer.
        """

        score = 0

        for row in range(
            SOURCE_HEADER_ROW + 1,
            worksheet.max_row + 1,
        ):
            for column in self._REQUIRED_COLUMNS:
                value = worksheet.cell(
                    row=row,
                    column=column,
                ).value

                if value is not None and str(value).strip():
                    score += 1

        return score
