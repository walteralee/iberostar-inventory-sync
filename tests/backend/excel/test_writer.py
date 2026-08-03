import math

import pytest
from openpyxl import Workbook

from excel.writer import ExcelWriter


@pytest.fixture
def worksheet():
    workbook = Workbook()
    return workbook.active


class TestWrite:
    def test_writes_into_an_empty_cell(self, worksheet):
        writer = ExcelWriter()

        writer.write(worksheet=worksheet, row=7, column=7, quantity=48)

        assert worksheet.cell(row=7, column=7).value == 48

    def test_accumulates_over_an_existing_numeric_value(self, worksheet):
        writer = ExcelWriter()
        worksheet.cell(row=7, column=7).value = 100

        writer.write(worksheet=worksheet, row=7, column=7, quantity=48)

        assert worksheet.cell(row=7, column=7).value == 148

    def test_two_deliveries_the_same_day_add_up(self, worksheet):
        # Este es el comportamiento que hace posible que dos entregas del
        # mismo producto en el mismo día (algo que sí ocurre en los Excel
        # reales de Economato) terminen sumadas en una sola celda.
        writer = ExcelWriter()

        writer.write(worksheet=worksheet, row=7, column=7, quantity=3024)
        writer.write(worksheet=worksheet, row=7, column=7, quantity=1512)

        assert worksheet.cell(row=7, column=7).value == 4536

    def test_treats_empty_string_cell_as_zero(self, worksheet):
        writer = ExcelWriter()
        worksheet.cell(row=7, column=7).value = ""

        writer.write(worksheet=worksheet, row=7, column=7, quantity=10)

        assert worksheet.cell(row=7, column=7).value == 10

    @pytest.mark.parametrize("row", [0, -1])
    def test_rejects_invalid_row(self, worksheet, row):
        writer = ExcelWriter()

        with pytest.raises(ValueError):
            writer.write(worksheet=worksheet, row=row, column=7, quantity=1)

    @pytest.mark.parametrize("column", [0, -1])
    def test_rejects_invalid_column(self, worksheet, column):
        writer = ExcelWriter()

        with pytest.raises(ValueError):
            writer.write(worksheet=worksheet, row=7, column=column, quantity=1)

    @pytest.mark.parametrize("quantity", [math.nan, math.inf, -math.inf])
    def test_rejects_non_finite_quantity(self, worksheet, quantity):
        writer = ExcelWriter()

        with pytest.raises(ValueError):
            writer.write(worksheet=worksheet, row=7, column=7, quantity=quantity)

    def test_rejects_boolean_quantity(self, worksheet):
        writer = ExcelWriter()

        with pytest.raises(ValueError):
            writer.write(worksheet=worksheet, row=7, column=7, quantity=True)

    def test_rejects_non_numeric_existing_cell_value(self, worksheet):
        writer = ExcelWriter()
        worksheet.cell(row=7, column=7).value = "no numérico"

        with pytest.raises(ValueError):
            writer.write(worksheet=worksheet, row=7, column=7, quantity=1)
