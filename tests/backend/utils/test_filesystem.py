"""
Pruebas del guardado atómico y las copias de seguridad con retención,
tanto para los Excel mensuales (Synchronizer) como para el Registry.

Importante: Synchronizer lee el directorio de backups desde
``config.settings.BACKUP_DIR``, que apunta a la carpeta real
``storage/backup`` del proyecto y no es inyectable por constructor. Para
no escribir ni borrar nada dentro del proyecto real durante los tests,
estos casos redirigen ``services.synchronizer.BACKUP_DIR`` a un directorio
temporal con ``monkeypatch`` en vez de dejar que apunte a la ruta real.
"""

from datetime import date

import pytest
from openpyxl import Workbook

from models.delivery import Delivery
from models.product import Product
from models.sales_point import SalesPoint
from services.registry import Registry
from services.synchronizer import Synchronizer


class _FakeRegistry:
    """Registry mínimo, solo para poder construir un Synchronizer."""

    def __init__(self):
        self.data = {}

    def exists(self, delivery):
        return False


def _make_delivery():
    return Delivery(
        sales_point=SalesPoint(name="Bar Piscina"),
        delivery_date=date(2026, 7, 1),
        products=[
            Product(code="1", name="A", format="UNIDAD", price=1.0, quantity=1.0)
        ],
    )


@pytest.fixture
def synchronizer(tmp_path, monkeypatch):
    monkeypatch.setattr("services.synchronizer.BACKUP_DIR", tmp_path / "backup")
    return Synchronizer(registry=_FakeRegistry())


class TestAtomicSaveWorkbook:
    def test_saves_workbook_to_target_path(self, synchronizer, tmp_path):
        workbook = Workbook()
        workbook.active["A1"] = "hola"
        target = tmp_path / "archivo.xlsx"

        synchronizer._atomic_save_workbook(workbook=workbook, target_path=target)

        assert target.is_file()

    def test_no_temporary_file_left_behind(self, synchronizer, tmp_path):
        workbook = Workbook()
        target = tmp_path / "archivo.xlsx"

        synchronizer._atomic_save_workbook(workbook=workbook, target_path=target)

        leftover_temp_files = list(tmp_path.glob(".*.tmp*"))
        assert leftover_temp_files == []

    def test_overwrites_existing_file_content(self, synchronizer, tmp_path):
        target = tmp_path / "archivo.xlsx"

        first = Workbook()
        first.active["A1"] = "version 1"
        synchronizer._atomic_save_workbook(workbook=first, target_path=target)

        second = Workbook()
        second.active["A1"] = "version 2"
        synchronizer._atomic_save_workbook(workbook=second, target_path=target)

        from openpyxl import load_workbook

        reloaded = load_workbook(target)
        assert reloaded.active["A1"].value == "version 2"
        reloaded.close()


class TestCreateBackupAndRetention:
    def test_raises_when_source_does_not_exist(self, synchronizer, tmp_path):
        with pytest.raises(FileNotFoundError):
            synchronizer._create_backup(
                source_path=tmp_path / "no_existe.xlsx",
                category="monthly",
            )

    def test_creates_a_copy_in_the_backup_directory(self, synchronizer, tmp_path):
        source = tmp_path / "Bar_Piscina_Julio_2026.xlsx"
        Workbook().save(source)

        backup_path = synchronizer._create_backup(
            source_path=source,
            category="monthly",
        )

        assert backup_path.is_file()
        assert backup_path.parent.name == "monthly"

    def test_keeps_only_the_configured_retention_count(self, synchronizer, tmp_path):
        source = tmp_path / "Bar_Piscina_Julio_2026.xlsx"
        Workbook().save(source)

        for _ in range(synchronizer._BACKUP_RETENTION_COUNT + 5):
            synchronizer._create_backup(source_path=source, category="monthly")

        backup_directory = synchronizer._create_backup(
            source_path=source,
            category="monthly",
        ).parent

        remaining = list(backup_directory.glob("Bar_Piscina_Julio_2026_*.xlsx"))
        assert len(remaining) == synchronizer._BACKUP_RETENTION_COUNT


class TestRegistryAtomicWriteAndRetention:
    def _make_registry(self, tmp_path):
        return Registry(
            registry_file=tmp_path / "imported_deliveries.json",
            backup_directory=tmp_path / "backup",
            acquire_lock=False,
        )

    def test_save_creates_the_registry_file(self, tmp_path):
        registry = self._make_registry(tmp_path)
        registry.register(_make_delivery())

        registry.save()

        assert (tmp_path / "imported_deliveries.json").is_file()
        registry.close()

    def test_save_backs_up_the_previous_content(self, tmp_path):
        registry = self._make_registry(tmp_path)
        registry.register(_make_delivery())
        registry.save()

        registry.mark_as_synchronized(_make_delivery())
        registry.save()

        backups = list((tmp_path / "backup").glob("imported_deliveries_*.json"))
        assert len(backups) >= 1
        registry.close()

    def test_keeps_only_the_configured_retention_count(self, tmp_path):
        registry = self._make_registry(tmp_path)

        for _ in range(registry._BACKUP_RETENTION_COUNT + 5):
            registry.save()

        backups = list((tmp_path / "backup").glob("imported_deliveries_*.json"))
        assert len(backups) == registry._BACKUP_RETENTION_COUNT
        registry.close()
