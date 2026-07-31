"""
Proyecto:
    Iberostar Inventory Synchronizer

Archivo:
    importer.py

Descripción:
    Servicio encargado de importar los movimientos contenidos
    en los archivos Excel de Economato.

    Las filas válidas se agrupan por fecha y punto de venta.
    Cada grupo se transforma en una entrega con sus productos.
"""

from __future__ import annotations

from datetime import date, datetime
from math import isclose, isfinite
from pathlib import Path
from tkinter import Tk, filedialog
from typing import Iterable
from zipfile import BadZipFile
import re
import unicodedata

from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from config.constants import (
    SALES_POINT_MAPPING,
    SOURCE_DATE_COLUMN,
    SOURCE_EXCEL_EXTENSION,
    SOURCE_FORMAT_COLUMN,
    SOURCE_GROUP_COLUMN,
    SOURCE_HEADER_ROW,
    SOURCE_PRICE_COLUMN,
    SOURCE_PRODUCT_CODE_COLUMN,
    SOURCE_PRODUCT_NAME_COLUMN,
    SOURCE_QUANTITY_COLUMN,
    SOURCE_SALES_POINT_COLUMN,
    SOURCE_SALES_POINT_PREFIX,
    VALID_PRODUCT_GROUPS,
)
from excel.source_reader import SourceReader
from models.delivery import Delivery
from models.product import Product
from models.sales_point import SalesPoint
from services.registry import Registry, RegistryConflictError
from utils.activity_log import log_incident
from utils.product_codes import normalize_product_code


class Importer:
    """
    Importa movimientos desde los Excel de Economato y los
    convierte en entregas agrupadas por fecha y punto de venta.

    El servicio:
        1. Selecciona o recibe los Excel de origen.
        2. Valida y procesa cada archivo de forma independiente.
        3. Agrupa productos por fecha y punto de venta.
        4. Construye las entregas.
        5. Consulta y actualiza el registro de sincronización.
    """

    _ZERO_TOLERANCE = 1e-9

    # Detecta un texto compuesto únicamente por grupos de millar en
    # formato español (p. ej. "1.234" o "12.345.678"), es decir, un
    # punto seguido de grupos de exactamente 3 dígitos y sin parte
    # decimal. Un texto que no encaje aquí (p. ej. "12.5") se sigue
    # tratando como decimal, igual que antes.
    _THOUSANDS_ONLY_PATTERN = re.compile(r"^-?\d{1,3}(\.\d{3})+$")

    _EXPECTED_FILE_ERRORS = (
        OSError,
        ValueError,
        TypeError,
        OverflowError,
        BadZipFile,
        InvalidFileException,
    )

    def __init__(
        self,
        registry: Registry,
        source_reader: SourceReader | None = None,
    ) -> None:
        """
        Inicializa los servicios necesarios para la importación.

        Args:
            registry: Registro compartido de entregas.
            source_reader: Lector de Excel opcional. Permite inyectar
                una implementación alternativa durante las pruebas.
        """

        self.registry = registry
        self.source_reader = source_reader or SourceReader()

        self._sales_point_mapping = {
            self._normalize_lookup_text(source_name): target_name
            for source_name, target_name in SALES_POINT_MAPPING.items()
        }

        self._valid_product_groups = {
            self._normalize_lookup_text(group) for group in VALID_PRODUCT_GROUPS
        }

        self._normalized_sales_point_prefix = self._normalize_lookup_text(
            SOURCE_SALES_POINT_PREFIX,
        )

    def run(
        self,
        excel_files: Iterable[Path | str] | Path | str | None = None,
    ) -> list[Delivery]:
        """
        Ejecuta el proceso completo de importación.

        Args:
            excel_files: Archivos que se deben procesar. Cuando no se
                proporcionan, se abre el explorador de archivos.

        Returns:
            Entregas nuevas o pendientes de sincronización.
        """

        self._print_section("1. SELECCIÓN DE ARCHIVOS EXCEL")

        if excel_files is None:
            print("Abriendo el explorador de archivos...")
            selected_files = self._select_excel_files()
        else:
            selected_files = self._normalize_file_paths(excel_files)

        if not selected_files:
            self._print_cancelled_selection()
            return []

        self._print_selected_files(selected_files)

        # La agrupación es global para permitir que varios archivos
        # aporten movimientos a una misma entrega.
        grouped_products: dict[
            tuple[date, str],
            dict[str, Product],
        ] = {}

        processed_file_count = 0
        file_error_count = 0

        total_rows_read = 0
        valid_row_count = 0
        ignored_group_count = 0
        ignored_sales_point_count = 0
        ignored_zero_quantity_count = 0
        ignored_duplicate_count = 0
        row_error_count = 0

        self._print_section("3. PROCESAMIENTO DE LOS EXCEL")

        for file_index, excel_file in enumerate(
            selected_files,
            start=1,
        ):
            worksheet: Worksheet | None = None

            # Cada archivo se procesa primero en una agrupación temporal.
            # Así, un error fatal no deja datos parciales del archivo dentro
            # de la importación global.
            file_grouped_products: dict[
                tuple[date, str],
                dict[str, Product],
            ] = {}

            file_rows_read = 0
            file_valid_rows = 0
            file_ignored_group_count = 0
            file_ignored_sales_point_count = 0
            file_ignored_zero_quantity_count = 0
            file_row_errors = 0

            try:
                self._print_file_header(
                    file_index=file_index,
                    total_files=len(selected_files),
                    excel_file=excel_file,
                )

                worksheet = self.source_reader.read(excel_file)
                self._validate_worksheet(worksheet)

                print(f"Hoja abierta     : {worksheet.title}")
                print(f"Filas detectadas : {worksheet.max_row}")
                print(f"Columnas         : {worksheet.max_column}")
                print("Proceso          : Leyendo movimientos...")
                print("-" * 100)

                for row_number in range(
                    SOURCE_HEADER_ROW + 1,
                    worksheet.max_row + 1,
                ):
                    row_values = self._read_source_row(
                        worksheet=worksheet,
                        row_number=row_number,
                    )

                    if self._is_empty_row(row_values):
                        continue

                    file_rows_read += 1

                    try:
                        product_group = self._normalize_lookup_text(
                            row_values["group"],
                        )

                        if product_group not in self._valid_product_groups:
                            file_ignored_group_count += 1
                            log_incident(
                                "Grupo de producto no admitido | "
                                f"archivo={excel_file.name} | "
                                f"fila={row_number} | "
                                f"grupo={row_values['group']!r}"
                            )
                            continue

                        sales_point = self._parse_sales_point(
                            row_values["sales_point"],
                        )

                        if sales_point is None:
                            file_ignored_sales_point_count += 1
                            log_incident(
                                "Punto de venta no reconocido | "
                                f"archivo={excel_file.name} | "
                                f"fila={row_number} | "
                                f"valor={row_values['sales_point']!r}"
                            )
                            continue

                        delivery_date = self._parse_date(
                            value=row_values["date"],
                            worksheet=worksheet,
                        )

                        product = self._parse_product(
                            code_value=row_values["code"],
                            name_value=row_values["name"],
                            format_value=row_values["format"],
                            price_value=row_values["price"],
                            quantity_value=row_values["quantity"],
                        )

                        if self._is_zero(product.quantity):
                            file_ignored_zero_quantity_count += 1
                            continue

                        self._add_product(
                            grouped_products=file_grouped_products,
                            delivery_date=delivery_date,
                            sales_point=sales_point,
                            product=product,
                            source_description=(
                                f"{excel_file.name}, fila {row_number}"
                            ),
                        )

                        file_valid_rows += 1

                    except (TypeError, ValueError, OverflowError) as error:
                        file_row_errors += 1

                        print(
                            f"Fila {row_number:05d} | "
                            f"IGNORADA | {type(error).__name__}: {error}"
                        )
                        log_incident(
                            "Fila con error de parseo | "
                            f"archivo={excel_file.name} | "
                            f"fila={row_number} | "
                            f"tipo={type(error).__name__} | "
                            f"motivo={error}"
                        )

                # Solo se incorporan los datos del archivo a la agrupación
                # global cuando el procesamiento termina correctamente.
                ignored_duplicate_count += self._merge_grouped_products(
                    target=grouped_products,
                    source=file_grouped_products,
                    source_description=excel_file.name,
                )

                processed_file_count += 1
                total_rows_read += file_rows_read
                valid_row_count += file_valid_rows
                ignored_group_count += file_ignored_group_count
                ignored_sales_point_count += file_ignored_sales_point_count
                ignored_zero_quantity_count += file_ignored_zero_quantity_count
                row_error_count += file_row_errors

                self._print_file_summary(
                    file_rows_read=file_rows_read,
                    file_valid_rows=file_valid_rows,
                    file_row_errors=file_row_errors,
                )

            except self._EXPECTED_FILE_ERRORS as error:
                file_error_count += 1

                self._print_expected_file_error(
                    file_index=file_index,
                    excel_file=excel_file,
                    error=error,
                )

            except Exception as error:
                # Un error inesperado suele señalar un fallo de programación.
                # Se muestra el contexto y se propaga para no ocultarlo.
                self._print_unexpected_file_error(
                    file_index=file_index,
                    excel_file=excel_file,
                    error=error,
                )
                raise

            finally:
                if worksheet is not None:
                    worksheet.parent.close()

        self._print_section("4. CONSTRUCCIÓN DE LAS ENTREGAS")

        deliveries = self._build_deliveries(grouped_products)

        print(f"Grupos encontrados  : {len(grouped_products)}")
        print(f"Entregas construidas: {len(deliveries)}")
        print("=" * 100)

        self._print_section("5. COMPROBACIÓN DEL REGISTRY")

        (
            imported_deliveries,
            imported_count,
            pending_count,
            existing_count,
            conflict_count,
        ) = self._filter_deliveries_with_registry(deliveries)

        self._print_section("6. GUARDADO DEL REGISTRY")

        self.registry.save()

        print("Estado: GUARDADO CORRECTAMENTE")
        print("=" * 100)

        self._print_import_summary(
            selected_file_count=len(selected_files),
            processed_file_count=processed_file_count,
            file_error_count=file_error_count,
            total_rows_read=total_rows_read,
            valid_row_count=valid_row_count,
            ignored_group_count=ignored_group_count,
            ignored_sales_point_count=ignored_sales_point_count,
            ignored_zero_quantity_count=ignored_zero_quantity_count,
            ignored_duplicate_count=ignored_duplicate_count,
            row_error_count=row_error_count,
            delivery_count=len(deliveries),
            imported_count=imported_count,
            pending_count=pending_count,
            existing_count=existing_count,
            conflict_count=conflict_count,
            synchronization_count=len(imported_deliveries),
        )

        return imported_deliveries

    # ======================================================
    # FILE PROCESSING
    # ======================================================

    def _select_excel_files(self) -> list[Path]:
        """
        Abre el explorador para seleccionar los Excel de origen.
        """

        root = Tk()

        try:
            root.withdraw()
            root.attributes("-topmost", True)

            files = filedialog.askopenfilenames(
                title="Seleccionar archivos Excel de Economato",
                filetypes=[
                    (
                        "Archivos Excel",
                        f"*{SOURCE_EXCEL_EXTENSION}",
                    ),
                ],
            )

        finally:
            root.destroy()

        return self._normalize_file_paths(files)

    def _normalize_file_paths(
        self,
        files: Iterable[Path | str] | Path | str,
    ) -> list[Path]:
        """
        Normaliza, elimina duplicados y ordena las rutas seleccionadas.
        """

        unique_files: dict[str, Path] = {}

        if isinstance(files, (str, Path)):
            files = [files]

        for file in files:
            path = Path(file).expanduser()
            key = str(path.resolve(strict=False)).casefold()
            unique_files.setdefault(key, path)

        return sorted(
            unique_files.values(),
            key=lambda path: str(path).casefold(),
        )

    def _validate_worksheet(
        self,
        worksheet: Worksheet,
    ) -> None:
        """
        Comprueba que la hoja tiene las dimensiones suficientes
        para poder procesarse.

        No valida el contenido de las cabeceras porque los informes
        reales de Economato pueden incluir nombres diferentes o
        algunas cabeceras vacías.
        """

        required_columns = (
            SOURCE_DATE_COLUMN,
            SOURCE_SALES_POINT_COLUMN,
            SOURCE_GROUP_COLUMN,
            SOURCE_PRODUCT_CODE_COLUMN,
            SOURCE_PRODUCT_NAME_COLUMN,
            SOURCE_FORMAT_COLUMN,
            SOURCE_QUANTITY_COLUMN,
            SOURCE_PRICE_COLUMN,
        )

        required_last_column = max(required_columns)

        if worksheet.max_column < required_last_column:
            raise ValueError(
                "El Excel no contiene todas las columnas necesarias. "
                f"Se requieren al menos {required_last_column} columnas "
                f"y se encontraron {worksheet.max_column}."
            )

        if worksheet.max_row <= SOURCE_HEADER_ROW:
            raise ValueError(
                "El Excel no contiene filas de datos después de la cabecera."
            )

    def _read_source_row(
        self,
        worksheet: Worksheet,
        row_number: int,
    ) -> dict[str, object]:
        """
        Lee exclusivamente las columnas utilizadas por el proyecto.
        """

        return {
            "date": worksheet.cell(
                row=row_number,
                column=SOURCE_DATE_COLUMN,
            ).value,
            "sales_point": worksheet.cell(
                row=row_number,
                column=SOURCE_SALES_POINT_COLUMN,
            ).value,
            "group": worksheet.cell(
                row=row_number,
                column=SOURCE_GROUP_COLUMN,
            ).value,
            "code": worksheet.cell(
                row=row_number,
                column=SOURCE_PRODUCT_CODE_COLUMN,
            ).value,
            "name": worksheet.cell(
                row=row_number,
                column=SOURCE_PRODUCT_NAME_COLUMN,
            ).value,
            "format": worksheet.cell(
                row=row_number,
                column=SOURCE_FORMAT_COLUMN,
            ).value,
            "quantity": worksheet.cell(
                row=row_number,
                column=SOURCE_QUANTITY_COLUMN,
            ).value,
            "price": worksheet.cell(
                row=row_number,
                column=SOURCE_PRICE_COLUMN,
            ).value,
        }

    def _is_empty_row(
        self,
        row_values: dict[str, object],
    ) -> bool:
        """
        Comprueba si todas las celdas relevantes están vacías.
        """

        return all(self._is_blank(value) for value in row_values.values())

    # ======================================================
    # VALUE PARSING
    # ======================================================

    def _parse_sales_point(
        self,
        value: object,
    ) -> SalesPoint | None:
        """
        Convierte el destino de Economato en un punto de venta interno.
        """

        normalized_value = self._normalize_lookup_text(value)

        if self._normalized_sales_point_prefix and normalized_value.startswith(
            self._normalized_sales_point_prefix,
        ):
            normalized_value = normalized_value[
                len(self._normalized_sales_point_prefix) :
            ].strip()

        mapped_name = self._sales_point_mapping.get(normalized_value)

        if mapped_name is None:
            return None

        return SalesPoint(name=mapped_name)

    def _parse_date(
        self,
        value: object,
        worksheet: Worksheet,
    ) -> date:
        """
        Convierte el valor de fecha del Excel en un objeto date.
        """

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numeric_value = float(value)

            if not isfinite(numeric_value):
                raise ValueError(f"La fecha numérica no es válida: {value}")

            parsed_value = from_excel(
                numeric_value,
                epoch=worksheet.parent.epoch,
            )

            if isinstance(parsed_value, datetime):
                return parsed_value.date()

            if isinstance(parsed_value, date):
                return parsed_value

        normalized_value = self._require_text(value, "fecha")

        accepted_formats = (
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%y",
            "%Y/%m/%d",
        )

        for date_format in accepted_formats:
            try:
                return datetime.strptime(
                    normalized_value,
                    date_format,
                ).date()
            except ValueError:
                continue

        raise ValueError(f"Formato de fecha no válido: {value}")

    def _parse_product(
        self,
        code_value: object,
        name_value: object,
        format_value: object,
        price_value: object,
        quantity_value: object,
    ) -> Product:
        """
        Construye un producto a partir de las celdas de una fila.
        """

        return Product(
            code=self._parse_product_code(code_value),
            name=self._require_text(
                name_value,
                "nombre del producto",
            ),
            format=self._require_text(
                format_value,
                "formato del producto",
            ),
            price=self._parse_number(price_value, "precio"),
            quantity=self._parse_number(quantity_value, "cantidad"),
        )

    def _parse_product_code(
        self,
        value: object,
    ) -> str:
        """
        Normaliza un código de producto numérico y conserva sus ceros
        iniciales cuando el Excel lo proporciona como texto.

        Esta regla coincide con la utilizada por ExcelFinder y
        ProductManager para reconocer las filas de productos dentro
        de las plantillas.
        """

        code = normalize_product_code(value)

        if code is None:
            raise ValueError(
                "El código del producto debe contener únicamente dígitos: "
                f"{value}"
            )

        return code

    def _parse_number(
        self,
        value: object,
        field_name: str,
    ) -> float:
        """
        Convierte un valor numérico del Excel a float y rechaza valores
        no finitos como NaN o infinito.
        """

        if isinstance(value, bool) or value is None:
            raise ValueError(f"El campo {field_name} no es numérico.")

        if isinstance(value, (int, float)):
            parsed_value = float(value)

            if not isfinite(parsed_value):
                raise ValueError(
                    f"El campo {field_name} contiene un número no válido: {value}"
                )

            return parsed_value

        normalized_value = str(value).strip()

        if not normalized_value:
            raise ValueError(f"El campo {field_name} está vacío.")

        normalized_value = (
            normalized_value.replace("\u00a0", "").replace("€", "").replace(" ", "")
        )

        is_parenthesized_negative = normalized_value.startswith(
            "("
        ) and normalized_value.endswith(")")

        if is_parenthesized_negative:
            normalized_value = normalized_value[1:-1]

        if "," in normalized_value and "." in normalized_value:
            if normalized_value.rfind(",") > normalized_value.rfind("."):
                normalized_value = normalized_value.replace(".", "")
                normalized_value = normalized_value.replace(",", ".")
            else:
                normalized_value = normalized_value.replace(",", "")

        elif "," in normalized_value:
            normalized_value = normalized_value.replace(",", ".")

        elif self._THOUSANDS_ONLY_PATTERN.fullmatch(normalized_value):
            # Sin coma decimal y con grupos de exactamente 3 dígitos:
            # es un separador de millar español, no un decimal.
            message = (
                f"El campo {field_name} = '{value}' se interpretó como "
                "separador de millar (formato español), no como decimal."
            )
            print(f"ADVERTENCIA | {message}")
            log_incident(message)
            normalized_value = normalized_value.replace(".", "")

        if is_parenthesized_negative:
            normalized_value = f"-{normalized_value}"

        try:
            parsed_value = float(normalized_value)
        except ValueError as error:
            raise ValueError(
                f"El campo {field_name} no es numérico: {value}"
            ) from error

        if not isfinite(parsed_value):
            raise ValueError(
                f"El campo {field_name} contiene un número no válido: {value}"
            )

        return parsed_value

    # ======================================================
    # GROUPING AND DELIVERY BUILDING
    # ======================================================

    def _add_product(
        self,
        grouped_products: dict[tuple[date, str], dict[str, Product]],
        delivery_date: date,
        sales_point: SalesPoint,
        product: Product,
        source_description: str = "origen desconocido",
    ) -> None:
        """
        Añade o acumula un producto dentro de su entrega.

        Cuando un mismo código presenta datos descriptivos distintos,
        se informa de la discrepancia y se conservan los datos más recientes.
        """

        delivery_key = (
            delivery_date,
            sales_point.name,
        )

        products = grouped_products.setdefault(delivery_key, {})
        existing_product = products.get(product.code)

        if existing_product is None:
            products[product.code] = product
            return

        differences: list[str] = []

        if existing_product.name != product.name:
            differences.append("nombre")

        if existing_product.format != product.format:
            differences.append("formato")

        if not isclose(
            existing_product.price,
            product.price,
            rel_tol=0.0,
            abs_tol=self._ZERO_TOLERANCE,
        ):
            differences.append("precio")

        if differences:
            print(
                "ADVERTENCIA | "
                f"Código {product.code}: cambian "
                f"{', '.join(differences)} en {source_description}. "
                "Se conservarán los datos más recientes."
            )

        existing_product.quantity += product.quantity
        existing_product.name = product.name
        existing_product.format = product.format
        existing_product.price = product.price

    def _merge_grouped_products(
        self,
        target: dict[tuple[date, str], dict[str, Product]],
        source: dict[tuple[date, str], dict[str, Product]],
        source_description: str,
    ) -> int:
        """
        Incorpora una agrupación procesada correctamente a la global.

        A diferencia de las filas dentro de un mismo Excel (donde un mismo
        código sí acumula cantidades, porque son movimientos distintos del
        mismo informe), entre Excel distintos NO se suman: los informes de
        Economato son acumulados y un Excel más reciente puede repetir
        movimientos de fecha y punto de venta que ya aportó un Excel
        anterior en esta misma ejecución. Sumarlos duplicaría cantidades.

        Por eso, si un código ya llegó de un Excel anterior para la misma
        fecha y punto de venta, se ignora la versión de este Excel y se
        conserva la ya incorporada.
        """

        ignored_duplicate_count = 0

        for (
            delivery_date,
            sales_point_name,
        ), products_by_code in source.items():
            target_products = target.setdefault(
                (delivery_date, sales_point_name),
                {},
            )

            for code, product in products_by_code.items():
                if code in target_products:
                    ignored_duplicate_count += 1

                    message = (
                        f"Código {code}: ya registrado para "
                        f"{delivery_date.strftime('%d/%m/%Y')} / "
                        f"{sales_point_name} en un Excel anterior. Se ignora "
                        f"en {source_description} para no duplicar cantidades."
                    )
                    print(f"ADVERTENCIA | {message}")
                    log_incident(message)
                    continue

                target_products[code] = product

        return ignored_duplicate_count

    def _build_deliveries(
        self,
        grouped_products: dict[tuple[date, str], dict[str, Product]],
    ) -> list[Delivery]:
        """
        Construye las entregas finales y elimina acumulaciones a cero.
        """

        deliveries: list[Delivery] = []

        for (
            delivery_date,
            sales_point_name,
        ), products_by_code in sorted(
            grouped_products.items(),
            key=lambda item: (
                item[0][0],
                item[0][1].casefold(),
            ),
        ):
            products = sorted(
                (
                    product
                    for product in products_by_code.values()
                    if not self._is_zero(product.quantity)
                ),
                key=lambda product: (
                    int(product.code),
                    product.code,
                ),
            )

            if not products:
                continue

            deliveries.append(
                Delivery(
                    sales_point=SalesPoint(name=sales_point_name),
                    delivery_date=delivery_date,
                    products=products,
                )
            )

        return deliveries

    # ======================================================
    # REGISTRY
    # ======================================================

    def _filter_deliveries_with_registry(
        self,
        deliveries: list[Delivery],
    ) -> tuple[list[Delivery], int, int, int, int]:
        """
        Clasifica las entregas como nuevas, pendientes o sincronizadas.

        Si una entrega coincide en fecha y punto de venta con una ya
        registrada pero con productos o cantidades diferentes, se trata
        como un conflicto: se omite solo esa entrega (no se sincroniza ni
        modifica el Registry) y el resto del proceso continúa con
        normalidad. El conflicto se deja constancia en el log de
        incidencias para revisión manual.
        """

        imported_deliveries: list[Delivery] = []
        imported_count = 0
        pending_count = 0
        existing_count = 0
        conflict_count = 0

        for delivery_index, delivery in enumerate(deliveries, start=1):
            print()
            print("-" * 100)
            print(f"ENTREGA {delivery_index:03d} " f"DE {len(deliveries):03d}")
            print("-" * 100)
            print("Fecha          : " f"{delivery.delivery_date.strftime('%d/%m/%Y')}")
            print(f"Punto de venta : {delivery.sales_point.name}")
            print(f"Productos      : {len(delivery.products)}")

            try:
                delivery_exists = self.registry.exists(delivery)
            except RegistryConflictError as error:
                conflict_count += 1

                print("Registry       : CONFLICTO DETECTADO")
                print(f"Detalle        : {error}")
                print("Resultado      : ENTREGA OMITIDA POR CONFLICTO")
                print("-" * 100)

                log_incident(
                    "Conflicto de Registry | "
                    f"fecha={delivery.delivery_date.strftime('%d/%m/%Y')} | "
                    f"punto_de_venta={delivery.sales_point.name} | "
                    f"detalle={error}"
                )
                continue

            if delivery_exists:
                if self.registry.is_synchronized(delivery):
                    existing_count += 1

                    print("Registry       : YA REGISTRADA")
                    print("Sincronización : COMPLETADA")
                    print("Resultado      : ENTREGA OMITIDA")
                    print("-" * 100)
                    continue

                pending_count += 1
                imported_deliveries.append(delivery)

                print("Registry       : YA REGISTRADA")
                print("Sincronización : PENDIENTE")
                print("Resultado      : ENTREGA RECUPERADA")
                print("-" * 100)
                continue

            self.registry.register(delivery)
            imported_deliveries.append(delivery)
            imported_count += 1

            print("Registry       : REGISTRADA")
            print("Sincronización : PENDIENTE")
            print("Resultado      : ENTREGA NUEVA")
            print("-" * 100)

        return (
            imported_deliveries,
            imported_count,
            pending_count,
            existing_count,
            conflict_count,
        )

    # ======================================================
    # GENERIC HELPERS
    # ======================================================

    def _require_text(
        self,
        value: object,
        field_name: str,
    ) -> str:
        """
        Devuelve un texto limpio y valida que no esté vacío.
        """

        if value is None:
            raise ValueError(f"El campo {field_name} está vacío.")

        normalized_value = " ".join(str(value).strip().split())

        if not normalized_value:
            raise ValueError(f"El campo {field_name} está vacío.")

        return normalized_value

    def _normalize_lookup_text(
        self,
        value: object,
    ) -> str:
        """
        Normaliza textos para comparaciones tolerantes a mayúsculas,
        espacios y acentos.
        """

        if value is None:
            return ""

        normalized_value = " ".join(str(value).strip().upper().split())

        normalized_value = unicodedata.normalize(
            "NFD",
            normalized_value,
        )

        return "".join(
            character
            for character in normalized_value
            if unicodedata.category(character) != "Mn"
        )

    def _is_blank(self, value: object) -> bool:
        """
        Comprueba si un valor está vacío o solo contiene espacios.
        """

        return value is None or str(value).strip() == ""

    def _is_zero(self, value: float) -> bool:
        """
        Compara cantidades con cero evitando residuos decimales mínimos.
        """

        return isclose(
            value,
            0.0,
            rel_tol=0.0,
            abs_tol=self._ZERO_TOLERANCE,
        )

    # ======================================================
    # CONSOLE OUTPUT
    # ======================================================

    def _print_section(self, title: str) -> None:
        print()
        print("=" * 100)
        print(title)
        print("=" * 100)

    def _print_cancelled_selection(self) -> None:
        print()
        print("!" * 100)
        print("SELECCIÓN CANCELADA")
        print("!" * 100)
        print("No se seleccionó ningún archivo Excel.")
        print("No hay archivos para importar.")
        print("!" * 100)

    def _print_selected_files(self, excel_files: list[Path]) -> None:
        self._print_section("2. LISTA DE EXCEL ENCONTRADOS")

        for index, excel_file in enumerate(excel_files, start=1):
            print(f"{index:03d} | {excel_file.name}")

        print("-" * 100)
        print(f"Total de Excel seleccionados: {len(excel_files)}")
        print("=" * 100)

    def _print_file_header(
        self,
        file_index: int,
        total_files: int,
        excel_file: Path,
    ) -> None:
        print()
        print("-" * 100)
        print(f"EXCEL {file_index:03d} DE {total_files:03d}" f" | {excel_file.name}")
        print("-" * 100)

    def _print_file_summary(
        self,
        file_rows_read: int,
        file_valid_rows: int,
        file_row_errors: int,
    ) -> None:
        print("-" * 100)
        print(f"Filas leídas     : {file_rows_read}")
        print(f"Filas válidas    : {file_valid_rows}")
        print(f"Errores de fila  : {file_row_errors}")
        print("Estado           : EXCEL PROCESADO")
        print("-" * 100)

    def _print_expected_file_error(
        self,
        file_index: int,
        excel_file: Path,
        error: Exception,
    ) -> None:
        print()
        print("!" * 100)
        print(f"ERROR DURANTE EL PROCESAMIENTO DEL EXCEL {file_index:03d}")
        print("!" * 100)
        print(f"Archivo : {excel_file.name}")
        print(f"Tipo    : {type(error).__name__}")
        print(f"Motivo  : {error}")
        print("Datos   : El archivo se descartó por completo")
        print("!" * 100)

    def _print_unexpected_file_error(
        self,
        file_index: int,
        excel_file: Path,
        error: Exception,
    ) -> None:
        print()
        print("!" * 100)
        print(f"ERROR INESPERADO EN EL EXCEL {file_index:03d}")
        print("!" * 100)
        print(f"Archivo : {excel_file.name}")
        print(f"Tipo    : {type(error).__name__}")
        print(f"Motivo  : {error}")
        print("Acción  : El error se propagará para no ocultar un fallo interno")
        print("!" * 100)

    def _print_import_summary(
        self,
        selected_file_count: int,
        processed_file_count: int,
        file_error_count: int,
        total_rows_read: int,
        valid_row_count: int,
        ignored_group_count: int,
        ignored_sales_point_count: int,
        ignored_zero_quantity_count: int,
        ignored_duplicate_count: int,
        row_error_count: int,
        delivery_count: int,
        imported_count: int,
        pending_count: int,
        existing_count: int,
        conflict_count: int,
        synchronization_count: int,
    ) -> None:
        self._print_section("7. RESUMEN DE IMPORTACIÓN")

        print(f"Excel seleccionados       : {selected_file_count}")
        print(f"Excel procesados          : {processed_file_count}")
        print(f"Excel con errores         : {file_error_count}")
        print("-" * 100)
        print(f"Filas leídas              : {total_rows_read}")
        print(f"Filas válidas             : {valid_row_count}")
        print(f"Grupos no admitidos       : {ignored_group_count}")
        print(f"Puntos de venta ignorados : {ignored_sales_point_count}")
        print(f"Cantidades a cero         : {ignored_zero_quantity_count}")
        print(f"Duplicados entre Excel    : {ignored_duplicate_count}")
        print(f"Filas con errores         : {row_error_count}")
        print("-" * 100)
        print(f"Entregas construidas      : {delivery_count}")
        print(f"Entregas nuevas           : {imported_count}")
        print(f"Entregas pendientes       : {pending_count}")
        print(f"Ya sincronizadas          : {existing_count}")
        print(f"En conflicto (omitidas)   : {conflict_count}")
        print(f"Entregas para sincronizar : {synchronization_count}")
        print("=" * 100)

        if conflict_count:
            print()
            print("!" * 100)
            print("ATENCIÓN: hay entregas omitidas por conflicto con el Registry.")
            print(
                "Revisa storage/logs/importacion_incidencias.log para "
                "identificar la fecha y el punto de venta afectados."
            )
            print("!" * 100)
