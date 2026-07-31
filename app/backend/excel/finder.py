"""
Proyecto:
    Iberostar Inventory Synchronizer

Archivo:
    finder.py

Descripción:
    Utilidades para localizar información dentro del Excel.

    Este módulo permite localizar la fila de un producto mediante su código
    y calcular la columna correspondiente al día del mes.
"""

from openpyxl.worksheet.worksheet import Worksheet

from config.constants import (
    PRODUCT_CODE_COLUMN,
    FIRST_PRODUCT_ROW,
    DAY_HEADER_ROW,
    FIRST_DAY_COLUMN,
)
from utils.product_codes import normalize_product_code


class ExcelFinder:
    """
    Localizador de información dentro del Excel.
    """

    def build_product_index(
        self,
        worksheet: Worksheet,
    ) -> dict[str, int]:
        """
        Construye un índice código → fila.

        Solo se indexan códigos válidos para mantener el
        comportamiento alineado con Importer y ProductManager.
        """

        index: dict[str, int] = {}

        for row in range(
            FIRST_PRODUCT_ROW,
            worksheet.max_row + 1,
        ):

            value = worksheet.cell(
                row=row,
                column=PRODUCT_CODE_COLUMN,
            ).value

            code = normalize_product_code(value)

            if code is None:
                continue

            # Si el código aparece varias veces se conserva
            # la primera aparición.
            index.setdefault(
                code,
                row,
            )

        print(f"Productos indexados : {len(index)}")
        print(f"Primera fila        : {FIRST_PRODUCT_ROW}")
        print(f"Última fila leída   : {worksheet.max_row}")
        print("Estado              : ÍNDICE CONSTRUIDO")
        print("-" * 100)

        return index

    def find_day_column(
        self,
        day: int,
    ) -> int:
        """
        Calcula la columna correspondiente al día del mes.
        """

        if not isinstance(day, int):
            raise TypeError("El día debe ser un número entero.")

        if not 1 <= day <= 31:
            raise ValueError(f"Día inválido: {day}")

        return FIRST_DAY_COLUMN + (day - 1)
