"""
Proyecto:
    Iberostar Inventory Synchronizer

Archivo:
    reader.py

Descripción:
    Lector de archivos Excel.

    Este módulo abre un archivo Excel y devuelve el libro junto
    con la hoja de trabajo principal utilizada por el proyecto.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config.constants import EXCEL_SHEET_NAME


class ExcelReader:
    """
    Lector de los Excel mensuales y de las plantillas.
    """

    def read(
        self,
        workbook_path: Path,
    ) -> tuple[Workbook, Worksheet]:
        """
        Abre un archivo Excel y devuelve el libro junto con
        la hoja configurada en EXCEL_SHEET_NAME.

        Args:
            workbook_path: Ruta del archivo Excel.

        Returns:
            Tupla formada por el libro y la hoja de trabajo.

        Raises:
            FileNotFoundError: Si el archivo no existe.
            ValueError: Si la ruta no es un archivo Excel válido
                o si no contiene la hoja requerida.
        """

        workbook_path = Path(
            workbook_path,
        )

        if not workbook_path.exists():

            raise FileNotFoundError(
                f"No se encontró el archivo Excel: {workbook_path}",
            )

        if not workbook_path.is_file():

            raise ValueError(
                f"La ruta no corresponde a un archivo: {workbook_path}",
            )

        if workbook_path.suffix.lower() != ".xlsx":

            raise ValueError(
                f"El archivo no tiene extensión .xlsx: {workbook_path.name}",
            )

        workbook = load_workbook(
            filename=workbook_path,
        )

        if EXCEL_SHEET_NAME not in workbook.sheetnames:

            available_sheets = ", ".join(
                workbook.sheetnames,
            )

            workbook.close()

            raise ValueError(
                f"No existe la hoja '{EXCEL_SHEET_NAME}' "
                f"en el archivo '{workbook_path.name}'. "
                f"Hojas disponibles: {available_sheets}",
            )

        worksheet = workbook[EXCEL_SHEET_NAME]

        return workbook, worksheet
