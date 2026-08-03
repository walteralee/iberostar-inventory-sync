"""
No existía ningún test para Importer pese a ser la puerta de entrada de
todos los datos: agrupación por fecha/punto de venta, parseo de números
en formato español y detección de duplicados. Todos los Excel de origen
se escriben en tmp_path y se pasan explícitamente a Importer.run(), para
no disparar nunca el diálogo de selección de archivos (tkinter).
"""

from datetime import date

import pytest
from openpyxl import Workbook

from services.importer import Importer
from services.registry import Registry

_HEADER_ROW = 1
_DATE_COL, _SALES_POINT_COL, _GROUP_COL = 1, 6, 7
_CODE_COL, _NAME_COL, _FORMAT_COL = 10, 12, 13
_QUANTITY_COL, _PRICE_COL = 16, 17


def _write_source_excel(path, rows):
    """
    rows: lista de tuplas
        (fecha, punto_de_venta, grupo, codigo, nombre, formato, cantidad, precio)
    """

    workbook = Workbook()
    worksheet = workbook.active

    for column in range(1, _PRICE_COL + 1):
        worksheet.cell(row=_HEADER_ROW, column=column).value = f"col{column}"

    for row_index, row in enumerate(rows, start=_HEADER_ROW + 1):
        (
            fecha,
            punto_de_venta,
            grupo,
            codigo,
            nombre,
            formato,
            cantidad,
            precio,
        ) = row

        worksheet.cell(row=row_index, column=_DATE_COL).value = fecha
        worksheet.cell(row=row_index, column=_SALES_POINT_COL).value = punto_de_venta
        worksheet.cell(row=row_index, column=_GROUP_COL).value = grupo
        worksheet.cell(row=row_index, column=_CODE_COL).value = codigo
        worksheet.cell(row=row_index, column=_NAME_COL).value = nombre
        worksheet.cell(row=row_index, column=_FORMAT_COL).value = formato
        worksheet.cell(row=row_index, column=_QUANTITY_COL).value = cantidad
        worksheet.cell(row=row_index, column=_PRICE_COL).value = precio

    workbook.save(path)
    return path


def _basic_row(**overrides):
    row = dict(
        fecha=date(2026, 7, 1),
        punto_de_venta="XAN - BAR PISCINA",
        grupo="BEBIDAS",
        codigo=7585,
        nombre="COCA COLA 33CL",
        formato="LATA",
        cantidad=48,
        precio=0.5,
    )
    row.update(overrides)
    return (
        row["fecha"],
        row["punto_de_venta"],
        row["grupo"],
        row["codigo"],
        row["nombre"],
        row["formato"],
        row["cantidad"],
        row["precio"],
    )


@pytest.fixture(autouse=True)
def _isolate_activity_log(tmp_path, monkeypatch):
    # log_incident() escribe en config.settings.LOGS_DIR, que no es
    # inyectable y por defecto apunta a storage/logs/ del proyecto real.
    # Se redirige aquí para que estos tests nunca dejen ruido en el log
    # real de incidencias de importación.
    monkeypatch.setattr("utils.activity_log.LOGS_DIR", tmp_path)
    monkeypatch.setattr(
        "utils.activity_log._LOG_FILE",
        tmp_path / "importacion_incidencias.log",
    )


@pytest.fixture
def registry(tmp_path):
    reg = Registry(
        registry_file=tmp_path / "imported_deliveries.json",
        backup_directory=tmp_path / "backup",
        acquire_lock=False,
    )
    yield reg
    reg.close()


@pytest.fixture
def importer(registry):
    return Importer(registry=registry)


class TestBasicGroupingAndMapping:
    def test_builds_one_delivery_per_date_and_sales_point(self, importer, tmp_path):
        path = _write_source_excel(tmp_path / "economato.xlsx", [_basic_row()])

        deliveries = importer.run([path])

        assert len(deliveries) == 1
        delivery = deliveries[0]
        assert delivery.delivery_date == date(2026, 7, 1)
        assert delivery.sales_point.name == "Bar_Piscina"
        assert len(delivery.products) == 1
        assert delivery.products[0].code == "7585"
        assert delivery.products[0].quantity == 48.0

    def test_unrecognized_sales_point_is_ignored(self, importer, tmp_path):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [_basic_row(punto_de_venta="XAN - PUNTO INVENTADO")],
        )

        deliveries = importer.run([path])

        assert deliveries == []
        assert importer.last_summary.ignored_sales_point_count == 1

    def test_invalid_product_group_is_ignored(self, importer, tmp_path):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [_basic_row(grupo="GRUPO NO VALIDO")],
        )

        deliveries = importer.run([path])

        assert deliveries == []
        assert importer.last_summary.ignored_group_count == 1

    def test_zero_quantity_row_is_ignored(self, importer, tmp_path):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [_basic_row(cantidad=0)],
        )

        deliveries = importer.run([path])

        assert deliveries == []
        assert importer.last_summary.ignored_zero_quantity_count == 1


class TestGroupingSumsWithinTheSameFile:
    def test_same_code_same_day_within_a_file_is_summed(self, importer, tmp_path):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [_basic_row(cantidad=48), _basic_row(cantidad=100)],
        )

        deliveries = importer.run([path])

        assert len(deliveries) == 1
        assert len(deliveries[0].products) == 1
        assert deliveries[0].products[0].quantity == 148.0

    def test_different_days_are_kept_as_separate_deliveries(self, importer, tmp_path):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [
                _basic_row(fecha=date(2026, 7, 1)),
                _basic_row(fecha=date(2026, 7, 2)),
            ],
        )

        deliveries = importer.run([path])

        assert len(deliveries) == 2
        assert {d.delivery_date for d in deliveries} == {
            date(2026, 7, 1),
            date(2026, 7, 2),
        }


class TestCrossFileDuplicates:
    def test_the_same_code_in_a_later_file_is_ignored_not_summed(
        self, importer, tmp_path
    ):
        # Documenta la decisión de diseño: entre Excel de Economato
        # distintos (informes acumulados) NO se suman, se conserva la
        # primera versión para no duplicar cantidades.
        first_file = _write_source_excel(
            tmp_path / "a_economato.xlsx", [_basic_row(cantidad=48)]
        )
        second_file = _write_source_excel(
            tmp_path / "b_economato.xlsx", [_basic_row(cantidad=999)]
        )

        deliveries = importer.run([first_file, second_file])

        assert len(deliveries) == 1
        assert deliveries[0].products[0].quantity == 48.0
        assert importer.last_summary.ignored_duplicate_count == 1


class TestSpanishNumberParsing:
    @pytest.mark.parametrize(
        "raw_value, expected",
        [
            ("48", 48.0),
            ("12,5", 12.5),
            ("1.234,56", 1234.56),
            ("1.234", 1234.0),
            ("(10)", -10.0),
        ],
    )
    def test_parses_spanish_formatted_quantities(
        self, importer, tmp_path, raw_value, expected
    ):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [_basic_row(cantidad=raw_value)],
        )

        deliveries = importer.run([path])

        assert deliveries[0].products[0].quantity == pytest.approx(expected)

    def test_thousands_only_format_logs_a_warning(self, importer, tmp_path):
        path = _write_source_excel(
            tmp_path / "economato.xlsx",
            [_basic_row(cantidad="1.234")],
        )

        importer.run([path])

        assert len(importer.last_summary.thousands_format_messages) == 1


class TestRegistryIntegration:
    def test_a_new_delivery_gets_registered_as_pending(self, importer, registry, tmp_path):
        path = _write_source_excel(tmp_path / "economato.xlsx", [_basic_row()])

        deliveries = importer.run([path])

        assert len(deliveries) == 1
        assert registry.exists(deliveries[0]) is True
        assert registry.is_synchronized(deliveries[0]) is False

    def test_reimporting_an_already_synchronized_delivery_is_skipped(
        self, importer, registry, tmp_path
    ):
        path = _write_source_excel(tmp_path / "economato.xlsx", [_basic_row()])

        first_pass = importer.run([path])
        registry.mark_as_synchronized(first_pass[0])
        registry.save()

        second_pass = importer.run([path])

        assert second_pass == []
        assert importer.last_summary.existing_count == 1

    def test_a_conflicting_reimport_is_reported_without_raising(
        self, importer, registry, tmp_path
    ):
        first_path = _write_source_excel(
            tmp_path / "economato1.xlsx", [_basic_row(cantidad=48)]
        )
        importer.run([first_path])

        second_path = _write_source_excel(
            tmp_path / "economato2.xlsx", [_basic_row(cantidad=999)]
        )
        deliveries = importer.run([second_path])

        assert deliveries == []
        assert importer.last_summary.conflict_count == 1
